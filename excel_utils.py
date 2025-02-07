from datetime import datetime
import io

def gerar_excel(clientes, status_vencimento):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        
        # Criar um novo workbook e selecionar a planilha ativa
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Clientes"

        # Definir estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
        
        # Adicionar cabeçalho
        headers = ['ID', 'Razão Social', 'CNPJ', 'Proprietário', 'Telefone', 'Data Vencimento', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Adicionar dados
        for row, cliente in enumerate(clientes, 2):
            ws.cell(row=row, column=1, value=cliente.id)
            ws.cell(row=row, column=2, value=cliente.razao_social)
            ws.cell(row=row, column=3, value=cliente.cnpj)
            ws.cell(row=row, column=4, value=cliente.proprietario)
            ws.cell(row=row, column=5, value=cliente.telefone)
            ws.cell(row=row, column=6, value=cliente.data_vencimento.strftime('%d/%m/%Y'))
            
            status = status_vencimento(cliente.data_vencimento)
            cell = ws.cell(row=row, column=7, value=status)
            
            # Definir cor do status
            if status == 'VENCIDO':
                cell.font = Font(color="FF0000")  # Vermelho
            elif status == 'PRÓXIMO AO VENCIMENTO':
                cell.font = Font(color="FFA500")  # Laranja
            else:
                cell.font = Font(color="008000")  # Verde

        # Ajustar largura das colunas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Salvar o arquivo em memória
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file
        
    except ImportError as e:
        raise ImportError("Biblioteca Excel não disponível") from e
    except Exception as e:
        raise Exception(f"Erro ao gerar Excel: {str(e)}") from e 