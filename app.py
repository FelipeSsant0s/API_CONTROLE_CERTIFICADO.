from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, abort, session
from flask_login import LoginManager, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
import os
import io
import openpyxl
import logging
import sys
import secrets
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from models import db, User, Certificado, Empresa
from werkzeug.utils import secure_filename
from api import api_bp
from werkzeug.exceptions import HTTPException
import traceback
import xml.etree.ElementTree as ET
import sqlalchemy
from sqlalchemy import text

# Configure logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

# Create Flask application
app = Flask(__name__)
logger.info('Initializing Flask application...')

# Configure application
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key')

# Database configuration
database_url = os.environ.get('DATABASE_URL')
if database_url:
    # Render.com usa 'postgres://', mas SQLAlchemy requer 'postgresql://'
    database_url = database_url.replace('postgres://', 'postgresql://')
    # Adicionar configuração SSL
    database_url += '?sslmode=require'
else:
    # Fallback para SQLite em desenvolvimento local
    database_url = 'sqlite:///certificados.db'

app.config['SQLALCHEMY_DATABASE_URI'] = database_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_size': 5,
    'max_overflow': 10,
    'pool_timeout': 30,
    'pool_recycle': 1800,
    'pool_pre_ping': True
}

# Initialize extensions
db.init_app(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Por favor, faça login para acessar esta página.'
login_manager.login_message_category = 'info'

# Initialize database and create admin user
def init_db():
    with app.app_context():
        try:
            logger.info('Starting database initialization...')
            # Forçar a criação das tabelas na ordem correta
            db.drop_all()  # Remove todas as tabelas existentes
            db.create_all()  # Recria as tabelas na ordem correta
            
            # Create default admin user if it doesn't exist
            admin_user = User.query.filter_by(username='admin').first()
            if not admin_user:
                admin = User(
                    username='admin',
                    email='admin@certificados.com',
                    name='Administrador'
                )
                admin.set_password('Admin@123')
                db.session.add(admin)
                db.session.commit()
                logger.info('Default admin user created')
            else:
                logger.info('Admin user already exists')
            
            logger.info('Database initialization completed successfully')
        except Exception as e:
            logger.error(f'Error in database initialization: {str(e)}')
            raise

# Call init_db to ensure database is initialized
init_db()

# Register blueprints
app.register_blueprint(api_bp, url_prefix='/api')

# Configuração para upload de arquivos
UPLOAD_FOLDER = 'uploads/xml'
ALLOWED_EXTENSIONS = {'xml'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Garantir que o diretório de upload existe
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

def determinar_status(data_validade):
    """Determina o status do certificado com base na data de validade"""
    try:
        hoje = datetime.now()
        if not isinstance(data_validade, datetime):
            logger.error(f'data_validade inválida: {data_validade} (tipo: {type(data_validade)})')
            return 'Erro'
            
        dias_para_vencer = (data_validade - hoje).days
        
        if dias_para_vencer < 0:
            return 'Expirado'
        elif dias_para_vencer <= 30:
            return 'Próximo ao Vencimento'
        else:
            return 'Válido'
    except Exception as e:
        logger.error(f'Erro ao determinar status: {str(e)}')
        return 'Erro'

# Authentication routes
@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(username=username).first()
        
        if user and user.check_password(password):
            login_user(user, remember=True)
            next_page = request.args.get('next')
            flash('Login realizado com sucesso!', 'success')
            return redirect(next_page or url_for('index'))
        else:
            flash('Usuário ou senha inválidos.', 'danger')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Logout realizado com sucesso!', 'success')
    return redirect(url_for('login'))

@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        username = request.form['username']
        email = request.form['email']
        name = request.form['name']
        password = request.form['password']
        
        if User.query.filter_by(username=username).first():
            flash('Este nome de usuário já está em uso.', 'danger')
            return redirect(url_for('register'))
        
        if User.query.filter_by(email=email).first():
            flash('Este e-mail já está em uso.', 'danger')
            return redirect(url_for('register'))
        
        user = User(username=username, email=email, name=name)
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        
        flash('Registro realizado com sucesso! Faça login para continuar.', 'success')
        return redirect(url_for('login'))
    
    return render_template('register.html')

# Protected routes
@app.route('/')
@login_required
def index():
    try:
        logger.info('Accessing index page')
        return render_template('index.html')
    except Exception as e:
        logger.error(f'Error in index route: {str(e)}')
        return render_template('500.html'), 500

@app.route('/certificados')
@login_required
def listar_certificados():
    search_query = request.args.get('search', '')
    data_inicial = request.args.get('data_inicial', '')
    data_final = request.args.get('data_final', '')
    
    query = Certificado.query
    
    if search_query:
        query = query.filter(
            (Certificado.nome_fantasia.ilike(f'%{search_query}%')) |
            (Certificado.razao_social.ilike(f'%{search_query}%')) |
            (Certificado.cnpj.ilike(f'%{search_query}%'))
        )
    
    if data_inicial:
        query = query.filter(Certificado.data_validade >= datetime.strptime(data_inicial, '%Y-%m-%d'))
    
    if data_final:
        query = query.filter(Certificado.data_validade <= datetime.strptime(data_final, '%Y-%m-%d'))
    
    certificados = query.order_by(Certificado.data_validade.asc()).all()
    
    return render_template('certificados.html', 
                         certificados=certificados,
                         search_query=search_query,
                         data_inicial=data_inicial,
                         data_final=data_final)

@app.route('/certificados/novo', methods=['GET', 'POST'])
@login_required
def novo_certificado():
    try:
        if request.method == 'POST':
            logger.info('Creating new certificate')
            try:
                # Log dos dados recebidos
                logger.info('Form data received:')
                for key, value in request.form.items():
                    logger.info(f'{key}: {value}')

                # Captura todos os campos do formulário
                razao_social = request.form.get('razao_social')
                nome_fantasia = request.form.get('nome_fantasia')
                cnpj = request.form.get('cnpj')
                telefone = request.form.get('telefone')
                data_validade_str = request.form.get('data_validade')
                observacoes = request.form.get('observacoes', '')  # Campo opcional

                logger.info('Validating required fields...')
                # Validação dos campos obrigatórios
                campos_obrigatorios = {
                    'Razão Social': razao_social,
                    'Nome Fantasia': nome_fantasia,
                    'CNPJ': cnpj,
                    'Telefone': telefone,
                    'Data de Validade': data_validade_str
                }

                campos_vazios = [campo for campo, valor in campos_obrigatorios.items() if not valor]
                if campos_vazios:
                    mensagem = f'Os seguintes campos são obrigatórios: {", ".join(campos_vazios)}'
                    logger.warning(f'Validation failed: {mensagem}')
                    flash(mensagem, 'danger')
                    return render_template('novo_certificado.html')

                logger.info('Converting date...')
                # Conversão da data
                try:
                    data_validade = datetime.strptime(data_validade_str, '%Y-%m-%d')
                except ValueError as e:
                    logger.error(f'Date conversion error: {str(e)}')
                    flash('Data de validade inválida. Use o formato YYYY-MM-DD.', 'danger')
                    return render_template('novo_certificado.html')

                logger.info('Checking for existing CNPJ...')
                # Verifica se já existe um certificado com este CNPJ para este usuário
                certificado_existente = Certificado.query.filter_by(
                    user_id=current_user.id,
                    cnpj=cnpj
                ).first()

                if certificado_existente:
                    logger.warning(f'Duplicate CNPJ found: {cnpj}')
                    flash('Já existe um certificado cadastrado com este CNPJ.', 'danger')
                    return render_template('novo_certificado.html')

                logger.info('Creating certificate object...')
                # Cria o certificado com status automático
                certificado = Certificado(
                    razao_social=razao_social,
                    nome_fantasia=nome_fantasia,
                    cnpj=cnpj,
                    telefone=telefone,
                    data_validade=data_validade,
                    observacoes=observacoes,
                    user_id=current_user.id
                )
                
                # Define o status automaticamente
                certificado.atualizar_status()
                logger.info(f'Certificate status set to: {certificado.status}')

                logger.info('Saving to database...')
                db.session.add(certificado)
                db.session.commit()
                logger.info(f'Certificate created successfully: {razao_social}')

                flash('Certificado criado com sucesso!', 'success')
                return redirect(url_for('listar_certificados'))

            except Exception as e:
                logger.error(f'Detailed error in form processing: {str(e)}', exc_info=True)
                db.session.rollback()
                flash(f'Erro ao processar os dados do formulário: {str(e)}', 'danger')
                return render_template('novo_certificado.html')

        return render_template('novo_certificado.html')
    except Exception as e:
        logger.error(f'General error in novo_certificado route: {str(e)}', exc_info=True)
        db.session.rollback()
        flash('Ocorreu um erro ao processar sua solicitação. Por favor, tente novamente.', 'danger')
        return render_template('novo_certificado.html')

@app.route('/certificados/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def editar_certificado(id):
    try:
        certificado = Certificado.query.get_or_404(id)
        if certificado.user_id != current_user.id:
            abort(403)
        
        if request.method == 'POST':
            logger.info(f'Updating certificate {id}')
            try:
                # Captura todos os campos do formulário
                razao_social = request.form.get('razao_social')
                nome_fantasia = request.form.get('nome_fantasia')
                novo_cnpj = request.form.get('cnpj')
                telefone = request.form.get('telefone')
                data_validade_str = request.form.get('data_validade')
                observacoes = request.form.get('observacoes')

                # Validação dos campos obrigatórios
                if not all([razao_social, nome_fantasia, novo_cnpj, telefone, data_validade_str]):
                    flash('Todos os campos obrigatórios devem ser preenchidos.', 'danger')
                    return render_template('editar_certificado.html', certificado=certificado)

                # Conversão da data
                try:
                    data_validade = datetime.strptime(data_validade_str, '%Y-%m-%d')
                except ValueError:
                    flash('Data de validade inválida.', 'danger')
                    return render_template('editar_certificado.html', certificado=certificado)

                # Verifica se o novo CNPJ já existe para outro certificado do mesmo usuário
                if novo_cnpj != certificado.cnpj:
                    certificado_existente = Certificado.query.filter_by(
                        user_id=current_user.id,
                        cnpj=novo_cnpj
                    ).first()

                    if certificado_existente:
                        flash('Já existe um certificado cadastrado com este CNPJ.', 'danger')
                        return render_template('editar_certificado.html', certificado=certificado)

                # Atualiza os dados do certificado
                certificado.razao_social = razao_social
                certificado.nome_fantasia = nome_fantasia
                certificado.cnpj = novo_cnpj
                certificado.telefone = telefone
                certificado.data_validade = data_validade
                certificado.observacoes = observacoes
                
                # Atualiza o status automaticamente
                certificado.atualizar_status()
                
                db.session.commit()
                logger.info(f'Certificate updated successfully: {certificado.razao_social}')
                
                flash('Certificado atualizado com sucesso!', 'success')
                return redirect(url_for('listar_certificados'))

            except Exception as e:
                logger.error(f'Error processing form data: {str(e)}')
                db.session.rollback()
                flash('Erro ao processar os dados do formulário. Por favor, verifique os campos e tente novamente.', 'danger')
                return render_template('editar_certificado.html', certificado=certificado)
        
        return render_template('editar_certificado.html', certificado=certificado)
    except Exception as e:
        logger.error(f'Error in editar_certificado route: {str(e)}')
        db.session.rollback()
        flash('Ocorreu um erro ao processar sua solicitação. Por favor, tente novamente.', 'danger')
        return render_template('editar_certificado.html', certificado=certificado)

@app.route('/certificados/<int:id>/deletar', methods=['POST'])
@login_required
def deletar_certificado(id):
    try:
        certificado = Certificado.query.get_or_404(id)
        if certificado.user_id != current_user.id:
            abort(403)
            
        razao_social = certificado.razao_social
        
        db.session.delete(certificado)
        db.session.commit()
        
        logger.info(f'Certificate deleted successfully: {razao_social}')
        flash('Certificado excluído com sucesso!', 'success')
        
        return redirect(url_for('listar_certificados'))
    except Exception as e:
        logger.error(f'Error in deletar_certificado route: {str(e)}')
        db.session.rollback()
        flash('Erro ao excluir certificado.', 'danger')
        return redirect(url_for('listar_certificados'))

@app.route('/certificados/exportar')
@login_required
def exportar_certificados():
    try:
        logger.info('Exporting certificates to Excel')
        certificados = Certificado.query.filter_by(user_id=current_user.id).all()
        # Atualiza o status antes de exportar
        for certificado in certificados:
            certificado.atualizar_status()
        db.session.commit()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Certificados"
        
        # Headers
        headers = ['ID', 'Razão Social', 'Nome Fantasia', 'CNPJ', 'Telefone', 'Data Emissão', 'Data Validade', 'Status', 'Observações']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Data
        for row, cert in enumerate(certificados, 2):
            ws.cell(row=row, column=1, value=cert.id)
            ws.cell(row=row, column=2, value=cert.razao_social)
            ws.cell(row=row, column=3, value=cert.nome_fantasia)
            ws.cell(row=row, column=4, value=cert.cnpj)
            ws.cell(row=row, column=5, value=cert.telefone)
            ws.cell(row=row, column=6, value=cert.data_emissao.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=7, value=cert.data_validade.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=8, value=cert.status)
            ws.cell(row=row, column=9, value=cert.observacoes)
        
        # Save to bytes
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        logger.info('Excel file generated successfully')
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='certificados.xlsx'
        )
    except Exception as e:
        logger.error(f'Error in exportar_certificados route: {str(e)}')
        return render_template('500.html'), 500

@app.route('/dashboard')
@login_required
def dashboard():
    try:
        # Atualizar status de todos os certificados
        Certificado.atualizar_status_todos()
        
        # Contar certificados por status
        stats = {
            'total': Certificado.query.count(),
            'validos': Certificado.query.filter_by(status='Válido').count(),
            'proximos': Certificado.query.filter_by(status='Próximo ao Vencimento').count(),
            'expirados': Certificado.query.filter_by(status='Expirado').count()
        }
        
        # Dados para o gráfico de pizza
        status_data = {
            'labels': ['Válidos', 'Próximos ao Vencimento', 'Expirados'],
            'data': [stats['validos'], stats['proximos'], stats['expirados']],
            'colors': ['#28a745', '#ffc107', '#dc3545']
        }
        
        # Preparar dados para o gráfico de linha (certificados por mês)
        hoje = datetime.utcnow()
        meses = []
        dados = []
        
        # Calcular os últimos 12 meses
        for i in range(12):
            mes = hoje - timedelta(days=30*i)
            meses.append(mes.strftime('%Y-%m'))
            
            # Contar certificados para este mês
            inicio = datetime.strptime(mes.strftime('%Y-%m'), '%Y-%m')
            fim = inicio + timedelta(days=30)
            count = Certificado.query.filter(
                Certificado.data_validade >= inicio,
                Certificado.data_validade < fim
            ).count()
            dados.append(count)
        
        # Ordenar meses e dados
        meses.reverse()
        dados.reverse()
        
        # Buscar próximos vencimentos
        vencimentos = Certificado.query.filter(
            Certificado.data_validade > hoje
        ).order_by(Certificado.data_validade).limit(5).all()
        
        # Calcular dias restantes para cada certificado
        for cert in vencimentos:
            cert.dias_restantes = (cert.data_validade - hoje).days
            
        # Buscar atividades recentes (últimos 5 certificados criados ou atualizados)
        atividades = Certificado.query.order_by(Certificado.data_emissao.desc()).limit(5).all()
        
        return render_template('dashboard.html', 
                             stats=stats,
                             status_data=status_data,
                             meses=meses,
                             dados=dados,
                             vencimentos=vencimentos,
                             atividades=atividades,
                             now=hoje)
                             
    except Exception as e:
        app.logger.error(f"Erro no dashboard: {str(e)}")
        app.logger.error(traceback.format_exc())
        return render_template('error.html', error=str(e)), 500

# Função para enviar email
def enviar_email(destinatario, assunto, corpo):
    try:
        # Configurações do email (você precisará definir estas variáveis de ambiente no Render.com)
        EMAIL_SENDER = os.environ.get('EMAIL_SENDER', 'seu-email@gmail.com')
        EMAIL_PASSWORD = os.environ.get('EMAIL_PASSWORD', 'sua-senha-app')
        
        # Criar mensagem
        msg = MIMEMultipart()
        msg['From'] = EMAIL_SENDER
        msg['To'] = destinatario
        msg['Subject'] = assunto
        
        # Adicionar corpo
        msg.attach(MIMEText(corpo, 'html'))
        
        # Conectar ao servidor SMTP do Gmail
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(EMAIL_SENDER, EMAIL_PASSWORD)
        
        # Enviar email
        server.send_message(msg)
        server.quit()
        return True
    except Exception as e:
        logger.error(f'Erro ao enviar email: {str(e)}')
        return False

# Modelo para códigos de recuperação de senha
class RecuperacaoSenha(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    codigo = db.Column(db.String(8), nullable=False)
    expiracao = db.Column(db.DateTime, nullable=False)
    usado = db.Column(db.Boolean, default=False)

# Rotas de Administração
@app.route('/admin')
@login_required
def admin():
    try:
        # Apenas admin pode acessar
        if current_user.username != 'admin':
            abort(403)
        
        users = User.query.all()
        return render_template('admin.html', users=users)
    except Exception as e:
        logger.error(f'Error in admin route: {str(e)}')
        return render_template('500.html'), 500

@app.route('/admin/alterar_senha/<int:user_id>', methods=['POST'])
@login_required
def admin_alterar_senha(user_id):
    try:
        # Apenas admin pode alterar
        if current_user.username != 'admin':
            abort(403)
        
        user = User.query.get_or_404(user_id)
        nova_senha = request.form.get('nova_senha')
        
        if not nova_senha:
            flash('A nova senha é obrigatória.', 'danger')
        else:
            user.set_password(nova_senha)
            db.session.commit()
            flash(f'Senha do usuário {user.username} alterada com sucesso!', 'success')
        
        return redirect(url_for('admin'))
    except Exception as e:
        logger.error(f'Error in admin_alterar_senha route: {str(e)}')
        db.session.rollback()
        flash('Erro ao alterar senha.', 'danger')
        return redirect(url_for('admin'))

# Rotas de Recuperação de Senha
@app.route('/recuperar_senha', methods=['GET', 'POST'])
def recuperar_senha():
    if request.method == 'POST':
        email = request.form.get('email')
        user = User.query.filter_by(email=email).first()
        
        if user:
            # Gerar código de verificação
            codigo = secrets.token_hex(4).upper()
            expiracao = datetime.utcnow() + timedelta(hours=1)
            
            # Salvar código no banco
            recuperacao = RecuperacaoSenha(
                user_id=user.id,
                codigo=codigo,
                expiracao=expiracao
            )
            db.session.add(recuperacao)
            db.session.commit()
            
            # Enviar email
            assunto = 'Recuperação de Senha - Gerenciador de Certificados'
            corpo = f"""
            <h2>Recuperação de Senha</h2>
            <p>Olá {user.name},</p>
            <p>Seu código de verificação é: <strong>{codigo}</strong></p>
            <p>Este código expira em 1 hora.</p>
            <p>Se você não solicitou a recuperação de senha, ignore este email.</p>
            """
            
            if enviar_email(user.email, assunto, corpo):
                flash('Código de verificação enviado para seu email!', 'success')
                return redirect(url_for('verificar_codigo', user_id=user.id))
            else:
                flash('Erro ao enviar email. Tente novamente.', 'danger')
        else:
            flash('Email não encontrado.', 'danger')
    
    return render_template('recuperar_senha.html')

@app.route('/verificar_codigo/<int:user_id>', methods=['GET', 'POST'])
def verificar_codigo(user_id):
    if request.method == 'POST':
        codigo = request.form.get('codigo')
        recuperacao = RecuperacaoSenha.query.filter_by(
            user_id=user_id,
            codigo=codigo,
            usado=False
        ).order_by(RecuperacaoSenha.expiracao.desc()).first()
        
        if recuperacao and recuperacao.expiracao > datetime.utcnow():
            recuperacao.usado = True
            db.session.commit()
            session['reset_user_id'] = user_id  # Armazenar ID do usuário para próxima etapa
            return redirect(url_for('nova_senha'))
        else:
            flash('Código inválido ou expirado.', 'danger')
    
    return render_template('verificar_codigo.html')

@app.route('/nova_senha', methods=['GET', 'POST'])
def nova_senha():
    user_id = session.get('reset_user_id')
    if not user_id:
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        senha = request.form.get('senha')
        confirmar_senha = request.form.get('confirmar_senha')
        
        if senha != confirmar_senha:
            flash('As senhas não coincidem.', 'danger')
        else:
            user = User.query.get(user_id)
            if user:
                user.set_password(senha)
                db.session.commit()
                session.pop('reset_user_id', None)  # Limpar sessão
                flash('Senha alterada com sucesso! Faça login com sua nova senha.', 'success')
                return redirect(url_for('login'))
    
    return render_template('nova_senha.html')

@app.route('/xml_upload', methods=['GET'])
@login_required
def xml_upload():
    return render_template('xml_upload.html')

@app.route('/upload_xml', methods=['POST'])
@login_required
def upload_xml():
    try:
        empresa_id = request.form.get('empresa_id')
        xml_file = request.files.get('xml_file')
        observacoes = request.form.get('observacoes', '')

        if not empresa_id or not xml_file:
            flash('ID da empresa e arquivo XML são obrigatórios', 'error')
            return redirect(url_for('xml_upload'))

        if not xml_file.filename.endswith('.xml'):
            flash('Apenas arquivos XML são permitidos', 'error')
            return redirect(url_for('xml_upload'))

        # Processar o XML
        xml_content = xml_file.read()
        # Aqui você pode adicionar a lógica para processar o XML
        # e extrair as informações do certificado

        # Criar um novo certificado
        novo_certificado = Certificado(
            empresa_id=empresa_id,
            observacoes=observacoes,
            # Adicione outros campos conforme necessário
        )
        db.session.add(novo_certificado)
        db.session.commit()

        flash('XML processado com sucesso!', 'success')
        return redirect(url_for('certificados'))

    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao processar XML: {str(e)}', 'error')
        return redirect(url_for('xml_upload'))

@app.route('/empresas')
@login_required
def empresas():
    empresas = Empresa.query.filter_by(user_id=current_user.id).all()
    return render_template('empresas.html', empresas=empresas)

@app.route('/empresas/nova', methods=['GET', 'POST'])
@login_required
def nova_empresa():
    if request.method == 'POST':
        try:
            empresa = Empresa(
                razao_social=request.form['razao_social'],
                nome_fantasia=request.form.get('nome_fantasia'),
                cnpj=request.form['cnpj'],
                endereco=request.form.get('endereco'),
                telefone=request.form.get('telefone'),
                email=request.form.get('email'),
                ativo=request.form.get('ativo') == 'on',
                url_integracao=Empresa.gerar_url_integracao(),
                user_id=current_user.id
            )
            db.session.add(empresa)
            db.session.commit()
            flash('Empresa cadastrada com sucesso!', 'success')
            return redirect(url_for('empresas'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao cadastrar empresa: {str(e)}', 'error')
            return redirect(url_for('nova_empresa'))
    return render_template('nova_empresa.html')

@app.route('/empresas/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def editar_empresa(id):
    empresa = Empresa.query.get_or_404(id)
    if empresa.user_id != current_user.id:
        abort(403)
    
    if request.method == 'POST':
        try:
            empresa.razao_social = request.form['razao_social']
            empresa.nome_fantasia = request.form.get('nome_fantasia')
            empresa.cnpj = request.form['cnpj']
            empresa.endereco = request.form.get('endereco')
            empresa.telefone = request.form.get('telefone')
            empresa.email = request.form.get('email')
            empresa.ativo = request.form.get('ativo') == 'on'
            
            db.session.commit()
            flash('Empresa atualizada com sucesso!', 'success')
            return redirect(url_for('empresas'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar empresa: {str(e)}', 'error')
            return redirect(url_for('editar_empresa', id=id))
    
    return render_template('editar_empresa.html', empresa=empresa)

@app.route('/empresas/<int:id>/deletar', methods=['POST'])
@login_required
def deletar_empresa(id):
    empresa = Empresa.query.get_or_404(id)
    if empresa.user_id != current_user.id:
        abort(403)
    
    try:
        db.session.delete(empresa)
        db.session.commit()
        flash('Empresa excluída com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir empresa: {str(e)}', 'error')
    
    return redirect(url_for('empresas'))

# Endpoint para integração com o monitor
@app.route('/api/integracao/<string:url_integracao>', methods=['POST'])
def receber_xml(url_integracao):
    try:
        empresa = Empresa.query.filter_by(url_integracao=url_integracao, ativo=True).first_or_404()
        
        if 'xml_file' not in request.files:
            return jsonify({'error': 'Nenhum arquivo enviado'}), 400
            
        xml_file = request.files['xml_file']
        if not xml_file.filename.endswith('.xml'):
            return jsonify({'error': 'Apenas arquivos XML são permitidos'}), 400
            
        # Processar o XML
        xml_content = xml_file.read()
        # Aqui você pode adicionar a lógica para processar o XML
        # e extrair as informações do certificado
        
        # Criar um novo certificado
        novo_certificado = Certificado(
            empresa_id=empresa.id,
            razao_social=empresa.razao_social,
            nome_fantasia=empresa.nome_fantasia,
            cnpj=empresa.cnpj,
            telefone=empresa.telefone,
            # Adicione outros campos conforme necessário
        )
        db.session.add(novo_certificado)
        db.session.commit()
        
        return jsonify({'message': 'XML processado com sucesso'}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

# Middleware para gerenciar a sessão do banco de dados
@app.teardown_appcontext
def shutdown_session(exception=None):
    db.session.remove()

@app.before_request
def before_request():
    try:
        # Verifica se a conexão está ativa usando text()
        db.session.execute(text('SELECT 1'))
    except Exception as e:
        logger.error(f"Erro na conexão com o banco: {str(e)}")
        db.session.rollback()
        db.session.close()
        db.session.remove()
        db.engine.dispose()

@app.after_request
def after_request(response):
    try:
        db.session.commit()
    except Exception as e:
        logger.error(f"Erro ao commitar transação: {str(e)}")
        db.session.rollback()
    finally:
        db.session.close()
    return response

# Modificar o manipulador de erros para lidar com erros de banco de dados
@app.errorhandler(Exception)
def handle_exception(e):
    logger.error(f"Erro não tratado: {str(e)}", exc_info=True)
    
    # Se for um erro de banco de dados, tenta reconectar
    if isinstance(e, (sqlalchemy.exc.OperationalError, sqlalchemy.exc.PendingRollbackError)):
        try:
            db.session.rollback()
            db.session.close()
            db.session.remove()
            db.engine.dispose()
            logger.info("Conexão com o banco de dados reiniciada após erro")
        except Exception as db_error:
            logger.error(f"Erro ao reiniciar conexão: {str(db_error)}")
    
    if isinstance(e, HTTPException):
        return render_template(f'{e.code}.html'), e.code
    
    return render_template('erro.html', error=str(e)), 500

@app.errorhandler(404)
def not_found_error(error):
    logger.error(f"Página não encontrada: {request.url}")
    return render_template('erro.html', error="Página não encontrada"), 404

@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    logger.error("Erro interno do servidor", exc_info=True)
    logger.error(f"Detalhes do erro: {str(error)}")
    logger.error(f"Stack trace: {error.__traceback__}")
    return render_template('erro.html', error=str(error)), 500

# Middleware para logging de requisições
@app.before_request
def log_request_info():
    logger.debug('Headers: %s', request.headers)
    logger.debug('Body: %s', request.get_data())
    logger.debug('URL: %s', request.url)
    logger.debug('Method: %s', request.method)
    logger.debug('Session: %s', session)
    logger.debug('User: %s', current_user if current_user.is_authenticated else 'Not authenticated')

@app.after_request
def log_response_info(response):
    logger.debug('Response Status: %s', response.status)
    logger.debug('Response Headers: %s', response.headers)
    return response

if __name__ == '__main__':
    app.run(debug=True) 